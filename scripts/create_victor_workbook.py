from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "artifacts"
OUTPUT_PATH = OUTPUT_DIR / "victor_wealth_triage_workbook_v1.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="203040")
SUBHEADER_FILL = PatternFill("solid", fgColor="D7B46A")
LIGHT_FILL = PatternFill("solid", fgColor="EEF3F8")
RED_FILL = PatternFill("solid", fgColor="FDE2E2")
GREEN_FILL = PatternFill("solid", fgColor="DFF4E5")
YELLOW_FILL = PatternFill("solid", fgColor="FFF4CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B7C3CF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_title(ws, cell):
    ws[cell].font = Font(size=14, bold=True)


def style_header_row(ws, row, end_col):
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BOX
        cell.alignment = WRAP


def style_key_value_row(ws, row):
    ws[f"A{row}"].font = BOLD
    ws[f"A{row}"].fill = LIGHT_FILL
    ws[f"A{row}"].border = BOX
    ws[f"B{row}"].border = BOX
    ws[f"C{row}"].border = BOX
    ws[f"B{row}"].alignment = WRAP
    ws[f"C{row}"].alignment = WRAP


def apply_yes_no_validation(ws, cell_range):
    validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def apply_list_validation(ws, cell_range, options):
    escaped = ",".join(options)
    validation = DataValidation(type="list", formula1=f'"{escaped}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def currency_cells(ws, *ranges):
    for cell_range in ranges:
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = '$#,##0'


def percent_cells(ws, *ranges):
    for cell_range in ranges:
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = '0.0%'


def build_readme(ws):
    ws.title = "README"
    ws["A1"] = "Victor Wealth Triage Workbook v1"
    style_title(ws, "A1")
    lines = [
        "Purpose: Structured intake and triage engine for a physician-specific financial assistant.",
        "Authority order:",
        "1. victor_sterling_financial_project_continuation_brief.md",
        "2. Brains/ constitutional source library",
        "3. This workbook and all derivative tools",
        "",
        "Usage:",
        "- Fill Intake, Cash Flow, Debts, Accounts, and Insurance first.",
        "- Review Dashboard for triage metrics and red flags.",
        "- Use Action Plan for ranked next steps.",
        "- Copy the output from Victor Prompt into your AI workflow when needed.",
        "",
        "Important:",
        "- Student-loan recommendations require current federal-rule verification.",
        "- Tax recommendations require current official-source verification.",
        "- Missing data should remain visible rather than guessed away.",
    ]
    for i, line in enumerate(lines, start=3):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 110
    ws.freeze_panes = "A3"


def build_intake(ws):
    ws["A1"] = "Intake"
    style_title(ws, "A1")
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    ws["C2"] = "Why it matters"
    style_header_row(ws, 2, 3)

    rows = [
        ("Client name", "", "Reference label only."),
        ("Age", "", "Affects timeline, risk capacity, and retirement runway."),
        ("State", "", "Tax and legal context."),
        ("Filing status", "", "Tax planning and loan strategy context."),
        ("Dependents", "", "Insurance need and liquidity pressure."),
        ("Training stage", "", "Resident, fellow, attending, or other changes almost everything."),
        ("Specialty", "", "Income trajectory and disability risk context."),
        ("Expected attending income", "", "Needed for training-to-attending transition planning."),
        ("Gross household income", "", "Base denominator for savings rate and planning."),
        ("Monthly take-home pay", "", "Real cash-flow planning depends on take-home, not fantasy gross."),
        ("Income stability", "", "Determines risk capacity and emergency-fund needs."),
        ("Near-term goals", "", "Prevents generic advice."),
        ("Notes", "", "Use for contract or family nuance."),
    ]
    for idx, (field, value, note) in enumerate(rows, start=3):
        ws[f"A{idx}"] = field
        ws[f"B{idx}"] = value
        ws[f"C{idx}"] = note
        style_key_value_row(ws, idx)

    apply_list_validation(ws, "B6", ["Single", "Married filing jointly", "Married filing separately", "Head of household", "Other / unsure"])
    apply_list_validation(ws, "B8", ["Resident", "Fellow", "Attending", "Private practice", "Other"])
    apply_list_validation(ws, "B13", ["Very stable", "Mostly stable", "Variable", "Unstable / at risk"])
    currency_cells(ws, "B10:B12")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 55
    ws.freeze_panes = "A3"


def build_cash_flow(ws):
    ws["A1"] = "Cash Flow"
    style_title(ws, "A1")
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    ws["C2"] = "Why it matters"
    style_header_row(ws, 2, 3)

    rows = [
        ("Monthly fixed expenses", "", "Housing, insurance, utilities, childcare, subscriptions."),
        ("Monthly variable expenses", "", "Food, gas, discretionary spending, and irregular leakage."),
        ("Monthly debt payments", "", "Required debt drag on cash flow."),
        ("Monthly savings and investing", "", "Core numerator for savings rate."),
        ("Cash buffer", "", "Liquid cash available now."),
        ("Target emergency-fund months", 6, "Set a realistic liquidity target."),
        ("Annual bonus or extra income", "", "Optional upside, not a planning crutch."),
    ]
    for idx, (field, value, note) in enumerate(rows, start=3):
        ws[f"A{idx}"] = field
        ws[f"B{idx}"] = value
        ws[f"C{idx}"] = note
        style_key_value_row(ws, idx)

    currency_cells(ws, "B3:B7", "B9:B9")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 58
    ws.freeze_panes = "A3"


def build_debts(ws):
    ws["A1"] = "Debts"
    style_title(ws, "A1")
    headers = [
        "Debt type",
        "Federal loan?",
        "Balance",
        "Interest rate",
        "Minimum monthly payment",
        "PSLF eligible?",
        "Qualifying payments",
        "Status / notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))

    for row in range(3, 15):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BOX
            ws.cell(row=row, column=col).alignment = WRAP

    apply_yes_no_validation(ws, "B3:B14")
    apply_yes_no_validation(ws, "F3:F14")
    currency_cells(ws, "C3:C14", "E3:E14")
    percent_cells(ws, "D3:D14")
    widths = {"A": 24, "B": 14, "C": 16, "D": 14, "E": 22, "F": 14, "G": 18, "H": 38}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def build_accounts(ws):
    ws["A1"] = "Accounts"
    style_title(ws, "A1")
    headers = [
        "Account type",
        "Balance",
        "Employer match %",
        "Current annual contribution",
        "Annual contribution limit",
        "Notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))

    for row in range(3, 15):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BOX
            ws.cell(row=row, column=col).alignment = WRAP

    currency_cells(ws, "B3:B14", "D3:E14")
    percent_cells(ws, "C3:C14")
    widths = {"A": 26, "B": 15, "C": 18, "D": 24, "E": 24, "F": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def build_insurance(ws):
    ws["A1"] = "Insurance"
    style_title(ws, "A1")
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    ws["C2"] = "Why it matters"
    style_header_row(ws, 2, 3)

    rows = [
        ("Own-occupation disability in force?", "", "Future income is often the physician's largest asset."),
        ("Employer disability coverage only?", "", "Group coverage may be inadequate or non-portable."),
        ("Term life in force?", "", "Relevant when dependents or shared debts exist."),
        ("Umbrella policy in force?", "", "Liability protection may matter as assets grow."),
        ("Malpractice tail issue identified?", "", "Contract and job-change risk."),
        ("Contract review completed?", "", "Bad employment terms can destroy wealth quietly."),
        ("Insurance notes", "", "Coverage details, riders, or concerns."),
    ]
    for idx, (field, value, note) in enumerate(rows, start=3):
        ws[f"A{idx}"] = field
        ws[f"B{idx}"] = value
        ws[f"C{idx}"] = note
        style_key_value_row(ws, idx)

    apply_yes_no_validation(ws, "B3:B8")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55
    ws.freeze_panes = "A3"


def build_dashboard(ws):
    ws["A1"] = "Dashboard"
    style_title(ws, "A1")
    ws["A2"] = "Metric"
    ws["B2"] = "Value"
    ws["C2"] = "Interpretation"
    style_header_row(ws, 2, 3)

    metrics = [
        ("Gross household income", "=Intake!B11", "Base income for planning."),
        ("Annual savings", "=IFERROR('Cash Flow'!B6*12,0)", "From current monthly savings/investing."),
        ("Savings rate", '=IFERROR(B4/B3,"")', "Below 15% is weak, 15-25% is decent, 25%+ is strong."),
        ("Monthly core spend", "=SUM('Cash Flow'!B3:B5)", "Core household cash outflow before optional nuance."),
        ("Emergency-fund months", '=IFERROR(\'Cash Flow\'!B7/B6,"")', "Liquidity buffer in months of core spend."),
        ("Total debt", "=SUM(Debts!C3:C14)", "All listed debt balances combined."),
        ("Federal student-loan balance", '=SUMIF(Debts!B3:B14,"Yes",Debts!C3:C14)', "Federal loans must be modeled before refinance talk."),
        ("High-interest debt flag", '=IF(COUNTIFS(Debts!B3:B14,"No",Debts!D3:D14,">=0.08",Debts!A3:A14,"<>Mortgage")>0,"Yes","No")', "Any expensive non-federal debt is a triage problem."),
        ("Employer match available", '=IF(COUNTIF(Accounts!C3:C14,">0")>0,"Yes","No")', "Free match should be captured before cleverness."),
        ("Disability protection gap", '=IF(AND(OR(Intake!B8="Resident",Intake!B8="Fellow",Intake!B8="Attending",Intake!B8="Private practice"),Insurance!B3<>"Yes"),"Review now","Looks covered or unknown")', "Physician income needs explicit protection."),
        ("Preliminary risk capacity", '=IF(OR(B7<3,Intake!B13="Unstable / at risk",B10="Yes"),"Constrained",IF(AND(B7>=6,OR(Intake!B13="Very stable",Intake!B13="Mostly stable"),B10="No"),"Moderate-to-High","Moderate"))', "Use the lower of risk capacity and risk tolerance in real planning."),
        ("Missing core inputs", '=COUNTBLANK(Intake!B4:B14)+COUNTBLANK(\'Cash Flow\'!B3:B8)', "Higher missingness means lower confidence in the plan."),
    ]
    start_row = 3
    for idx, (name, formula, note) in enumerate(metrics, start=start_row):
        ws[f"A{idx}"] = name
        ws[f"B{idx}"] = formula
        ws[f"C{idx}"] = note
        style_key_value_row(ws, idx)

    currency_cells(ws, "B3:B4", "B6:B6", "B8:B9")
    percent_cells(ws, "B5:B5")
    ws["E2"] = "Read This"
    ws["E2"].fill = SUBHEADER_FILL
    ws["E2"].font = BOLD
    ws["E2"].border = BOX
    notes = [
        "This dashboard is triage, not final advice.",
        "Missing data should reduce confidence, not trigger invented precision.",
        "Federal student-loan choices require current official-rule verification.",
        "Tax optimization requires live current-law review.",
    ]
    for i, note in enumerate(notes, start=3):
        ws[f"E{i}"] = note
        ws[f"E{i}"].border = BOX
        ws[f"E{i}"].alignment = WRAP

    ws["B5"].fill = YELLOW_FILL
    ws["B7"].fill = YELLOW_FILL
    ws["B10"].fill = RED_FILL
    ws["B11"].fill = YELLOW_FILL
    ws["B12"].fill = RED_FILL
    ws["B13"].fill = GREEN_FILL
    ws["B14"].fill = YELLOW_FILL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["E"].width = 48
    ws.freeze_panes = "A3"


def build_action_plan(ws):
    ws["A1"] = "Action Plan"
    style_title(ws, "A1")
    headers = ["Priority", "Category", "Victor verdict", "Reason", "Immediate next step"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))

    rows = [
        (
            1,
            "Liquidity",
            '=IF(Dashboard!B7<1,"Critical: liquidity is dangerously thin",IF(Dashboard!B7<3,"Weak: build cash before optimization","Acceptable for now"))',
            "Victor protects against ruin first.",
            '=IF(Dashboard!B7<3,"Build starter cash reserve and stop optional complexity","Maintain or continue toward target reserve")',
        ),
        (
            2,
            "High-interest debt",
            '=IF(Dashboard!B10="Yes","Attack this now","No obvious high-interest debt emergency detected")',
            "Guaranteed interest avoidance usually beats speculative expected return.",
            '=IF(Dashboard!B10="Yes","Direct surplus cash toward expensive debt payoff","Proceed to next priority")',
        ),
        (
            3,
            "Disability insurance",
            '=IF(Dashboard!B12="Review now","Protection gap identified","No immediate gap detected or details missing")',
            "For physicians, future income is often the largest asset.",
            '=IF(Dashboard!B12="Review now","Get own-occupation disability quotes and review riders","Confirm coverage details and portability")',
        ),
        (
            4,
            "Employer match",
            '=IF(Dashboard!B11="Yes","Capture full available match","No match identified or data missing")',
            "Match dollars come before clever optimization.",
            '=IF(Dashboard!B11="Yes","Verify contribution rate reaches the full match threshold","Confirm whether a match exists")',
        ),
        (
            5,
            "Federal student loans",
            '=IF(Dashboard!B9>0,"Model PSLF/IDR before any refinance talk","No federal student-loan balance detected")',
            "Never casually refinance away federal options.",
            '=IF(Dashboard!B9>0,"Gather loan type, servicer, PSLF count, repayment plan, and spouse context","Proceed only after confirming loan details")',
        ),
        (
            6,
            "Tax-advantaged investing",
            '=IF(AND(Dashboard!B7>=3,Dashboard!B10="No"),"Advance once the foundation is stable","Do not optimize before stabilizing the base")',
            "Optimization should follow stability, not replace it.",
            "Review 401k/403b, HSA, IRA, and other available account space.",
        ),
        (
            7,
            "Long-term investing",
            "Default to low-cost diversified index funds once the base is structurally sound.",
            "Complexity must earn its place.",
            "Write or refine a target allocation and contribution plan.",
        ),
    ]

    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            ws.cell(row=row_idx, column=col_idx).border = BOX
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP

    widths = {"A": 10, "B": 22, "C": 42, "D": 44, "E": 52}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def build_prompt(ws):
    ws["A1"] = "Victor Prompt"
    style_title(ws, "A1")
    ws["A2"] = "Copy or adapt the text below when you want Victor to analyze the workbook."
    ws["A4"] = (
        '=TEXTJOIN(CHAR(10),TRUE,'
        '"You are Victor Sterling, a physician-specific financial strategist.",'
        '"Use the workbook data below. Label assumptions instead of faking certainty.",'
        '"",'
        '"CLIENT SNAPSHOT",'
        '"- Client name: "&Intake!B3,'
        '"- Age: "&Intake!B4,'
        '"- State: "&Intake!B5,'
        '"- Filing status: "&Intake!B6,'
        '"- Dependents: "&Intake!B7,'
        '"- Training stage: "&Intake!B8,'
        '"- Specialty: "&Intake!B9,'
        '"- Expected attending income: "&TEXT(Intake!B10,"$#,##0"),'
        '"- Gross household income: "&TEXT(Intake!B11,"$#,##0"),'
        '"- Monthly take-home pay: "&TEXT(Intake!B12,"$#,##0"),'
        '"- Income stability: "&Intake!B13,'
        '"- Near-term goals: "&Intake!B14,'
        '"",'
        '"KEY METRICS",'
        '"- Annual savings: "&TEXT(Dashboard!B4,"$#,##0"),'
        '"- Savings rate: "&TEXT(Dashboard!B5,"0.0%"),'
        '"- Monthly core spend: "&TEXT(Dashboard!B6,"$#,##0"),'
        '"- Emergency-fund months: "&TEXT(Dashboard!B7,"0.0"),'
        '"- Total debt: "&TEXT(Dashboard!B8,"$#,##0"),'
        '"- Federal student-loan balance: "&TEXT(Dashboard!B9,"$#,##0"),'
        '"- High-interest debt flag: "&Dashboard!B10,'
        '"- Employer match available: "&Dashboard!B11,'
        '"- Disability protection gap: "&Dashboard!B12,'
        '"- Preliminary risk capacity: "&Dashboard!B13,'
        '"- Missing core inputs: "&Dashboard!B14,'
        '"",'
        '"TASK",'
        '"Produce: executive summary, diagnosis, assumptions, risk assessment, debt strategy, insurance strategy, retirement strategy, portfolio strategy, tax notes, lifestyle-inflation controls, and action timeline.",'
        '"Verify current law before making tax or student-loan recommendations."'
        ')'
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A4"].border = BOX
    ws["A4"].fill = LIGHT_FILL
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[4].height = 420
    ws.freeze_panes = "A4"


def build_workbook():
    wb = Workbook()
    build_readme(wb.active)
    build_intake(wb.create_sheet("Intake"))
    build_cash_flow(wb.create_sheet("Cash Flow"))
    build_debts(wb.create_sheet("Debts"))
    build_accounts(wb.create_sheet("Accounts"))
    build_insurance(wb.create_sheet("Insurance"))
    build_dashboard(wb.create_sheet("Dashboard"))
    build_action_plan(wb.create_sheet("Action Plan"))
    build_prompt(wb.create_sheet("Victor Prompt"))

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build_workbook()
    print(path)
